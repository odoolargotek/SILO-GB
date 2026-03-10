# -*- coding: utf-8 -*-

# LT Brigade Module - Mejoras Odoo 18
# Largotek SRL - Juan Luis Garvía - www.largotek.com
# License: LGPL-3.0 (https://www.gnu.org/licenses/lgpl-3.0.html)

from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError
import base64
import io
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Please install openpyxl: pip install openpyxl")


class GBBrigade(models.Model):
    _name = "gb.brigade"
    _description = "Global Brigades - Chapter / Brigade"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"

    # =========================
    # CAMPO ACTIVE (para archivar)
    # =========================
    active = fields.Boolean(
        string="Active",
        default=True,
        help="If unchecked, this brigade will be archived and hidden from main views.",
        tracking=True,
    )

    # =========================
    # COMPANY (Multi-company support)
    # =========================
    company_id = fields.Many2one(
        'res.company',
        string='Company',
        default=lambda self: self.env.company,
        required=True,
        tracking=True,
    )

    # =========================
    # IDENTIFICACIÓN BÁSICA
    # =========================
    external_brigade_code = fields.Char(
        string="Brigade Code",
        help="External reference / code from CRM or other system.",
        tracking=True,
    )

    brigade_code = fields.Char(
        string="Internal Code",
        readonly=True,
        copy=False,
        default="/",
        tracking=True,
    )

    name = fields.Char(string="Chapter Name", required=True, tracking=True)
    arrival_date = fields.Date(string="Arrival Date", tracking=True)
    departure_date = fields.Date(string="Departure Date", tracking=True)

    # =========================
    # ESTADO OPERATIVO
    # =========================
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("planned", "Planned"),
            ("ready", "Ready"),
            ("in_field", "In Field"),
            ("completed", "Completed"),
            ("archived", "Archived"),
        ],
        string="Status",
        default="draft",
        required=True,
        help="Operational state of the Brigade",
        tracking=True,
    )

    brigade_type = fields.Selection(
        [
            ("onsite", "In Person"),
            ("virtual", "Virtual"),
        ],
        string="Brigade Type",
        default="onsite",
        required=True,
        help="If set to Virtual, logistics sections should not be used.",
        tracking=True,
    )

    brigade_restriction = fields.Selection(
        [
            ("no_restrictions", "No Restrictions"),
            ("darien_golfo", "Darien & Golfo de Mosquito"),
            ("este_darien", "Este y Darien"),
            ("solo_darien", "Solo Darien"),
            ("otros", "Otros"),
        ],
        string="Restrictions",
        help="Geographic / operational restrictions.",
        tracking=True,
    )

    brigade_program = fields.Many2one(
        'gb.brigade.program.type',
        string="Official Program",
        help="Main official program.",
        tracking=True,
    )

    # =========================
    # LT ITINERARY SIMPLE
    # =========================
    lt_itinerary_link = fields.Char(
        string="Itinerary Link",
        help="Pegar link de GDrive o cualquier URL.",
    )

    lt_itinerary_locked = fields.Boolean(
        string="Bloquear Link",
        default=False,
    )

    lt_itinerary_url = fields.Char(
        string="Itinerary URL",
        compute="_compute_lt_itinerary_url",
        store=True,
        readonly=True,
    )

    # =========================
    # BUSINESS / TIER
    # =========================
    business_client_id = fields.Many2one(
        "res.partner",
        string="Business Client",
        help="Client when program is Business.",
        tracking=True,
    )

    business_profile_link = fields.Char(
        string="Business Profile Link",
        related="business_client_id.business_profile_link",
        readonly=True,
    )

    @api.onchange("brigade_program")
    def _onchange_brigade_program_business_client(self):
        for record in self:
            if not record.brigade_program or record.brigade_program.name.lower() != 'business':
                record.business_client_id = False

    brigade_tier = fields.Selection(
        [
            ("sustainable", "Sustainable"),
            ("empowered", "Empowered"),
            ("scaled", "Scaled"),
        ],
        string="Brigade Tier",
        help="Tier: Sustainable(14-25), Empowered(26-39), Scaled(40+).",
        tracking=True,
    )

    # =========================
    # CONTADORES UNIFICADOS
    # =========================
    volunteer_count = fields.Integer(
        string="Volunteers", compute="_compute_counts"
    )
    program_count = fields.Integer(
        string="Programs", compute="_compute_counts"
    )
    activity_count = fields.Integer(
        string="Activities", compute="_compute_counts"
    )
    transport_count = fields.Integer(
        string="Transports", compute="_compute_counts"
    )

    # =========================
    # KPI COUNTS (simplified)
    # =========================
    staff_count = fields.Integer(
        string="Staff Count",
        compute="_compute_kpi_counts",
        store=False,
        help="Total number of staff members assigned.",
    )
    total_participants = fields.Integer(
        string="Total Participants",
        compute="_compute_kpi_counts",
        store=False,
        help="Total participants (Roster + Staff).",
    )

    university_logo = fields.Image(string="University Logo")
    
    # =========================
    # CONTACTOS MULTIPLES (Many2many)
    # =========================
    success_advisor_ids = fields.Many2many(
        "res.partner",
        "gb_brigade_success_advisor_rel",
        "brigade_id",
        "partner_id",
        string="Success Advisor",
        tracking=True,
        help="Success Advisors assigned to this brigade (multiple allowed)",
    )
    
    coordinator_ids = fields.Many2many(
        "res.partner",
        "gb_brigade_coordinator_rel",
        "brigade_id",
        "partner_id",
        string="Lead Coordinator",
        tracking=True,
        help="Lead Coordinators assigned to this brigade (multiple allowed)",
    )
    
    program_associate_ids = fields.Many2many(
        "res.partner",
        "gb_brigade_program_associate_rel",
        "brigade_id",
        "partner_id",
        string="Program Advisor",
        tracking=True,
        help="Program Advisors assigned to this brigade (multiple allowed)",
    )
    
    sending_organization_ids = fields.Many2many(
        "res.partner",
        "gb_brigade_sending_org_rel",
        "brigade_id",
        "partner_id",
        string="Sending Organization",
        tracking=True,
        help="Sending Organizations for this brigade (multiple allowed)",
    )

    chapter_leader_ids = fields.Many2many(
        "gb.brigade.roster",
        "gb_brigade_chapter_leader_rel",
        "brigade_id",
        "roster_id",
        string="Chapter Leader(s)",
    )
    chapter_president_faculty_ids = fields.Many2many(
        "gb.brigade.roster",
        "gb_brigade_chapter_president_faculty_rel",
        "brigade_id",
        "roster_id",
        string="Chapter President / Faculty",
    )
    professor_chaperone_ids = fields.Many2many(
        "gb.brigade.roster",
        "gb_brigade_professor_chaperone_rel",
        "brigade_id",
        "roster_id",
        string="Professor / Chaperone",
    )

    extra_info = fields.Text(string="Additional Information")

    # =========================
    # ONE2MANY RELACIONES
    # =========================
    program_line_ids = fields.One2many(
        "gb.brigade.program", "brigade_id", string="Programs"
    )
    roster_ids = fields.One2many(
        "gb.brigade.roster", "brigade_id", string="Roster"
    )
    arrival_ids = fields.One2many(
        "gb.brigade.arrival", "brigade_id", string="Arrivals"
    )
    departure_ids = fields.One2many(
        "gb.brigade.departure", "brigade_id", string="Departures"
    )
    staff_ids = fields.One2many(
        "gb.brigade.staff", "brigade_id", string="Temp Staff"
    )
    # RENAMED: activity_ids -> brigade_activity_ids to avoid conflict with mail.activity.mixin
    brigade_activity_ids = fields.One2many(
        "gb.brigade.activity", "brigade_id", string="Activities"
    )
    hotel_booking_ids = fields.One2many(
        "gb.brigade.hotel.booking", "brigade_id", string="Hotel"
    )
    transport_ids = fields.One2many(
        "gb.brigade.transport", "brigade_id", string="Transport"
    )

    # =========================
    # S.A. NOTIFICATION
    # =========================
    sa_pending_count = fields.Integer(
        string="S.A. Pending",
        compute="_compute_sa_pending_count",
        store=False,
    )

    _sql_constraints = [
        (
            "chapter_code_uniq",
            "unique(brigade_code)",
            "Brigade Code must be unique!",
        ),
    ]

    # =========================
    # COMPUTES / CONSTRAINS
    # =========================
    @api.depends("roster_ids", "program_line_ids", "brigade_activity_ids", "transport_ids")
    def _compute_counts(self):
        for rec in self:
            rec.volunteer_count = len(rec.roster_ids)
            rec.program_count = len(rec.program_line_ids)
            rec.activity_count = len(rec.brigade_activity_ids)
            rec.transport_count = len(rec.transport_ids)

    @api.depends("volunteer_count", "staff_ids")
    def _compute_kpi_counts(self):
        """Compute simplified KPI counts for brigade."""
        for rec in self:
            staff_total = len(rec.staff_ids)
            rec.staff_count = staff_total
            rec.total_participants = rec.volunteer_count + staff_total

    @api.depends("roster_ids.sa", "roster_ids.sa_notified")
    def _compute_sa_pending_count(self):
        for rec in self:
            rec.sa_pending_count = len(
                rec.roster_ids.filtered(
                    lambda r: r.sa and not r.sa_notified and r.partner_id.gb_emergency_contact_id
                )
            )

    @api.constrains(
        "brigade_type",
        "transport_ids",
        "hotel_booking_ids",
        "arrival_ids",
        "departure_ids",
    )
    def _check_virtual_no_logistics(self):
        for rec in self:
            if rec.brigade_type == "virtual" and (
                rec.transport_ids
                or rec.hotel_booking_ids
                or rec.arrival_ids
                or rec.departure_ids
            ):
                raise ValidationError(
                    _("Virtual brigades cannot have logistics records.")
                )

    # =========================
    # ITINERARY URL (compute)
    # =========================
    @api.depends("lt_itinerary_link")
    def _compute_lt_itinerary_url(self):
        for rec in self:
            if rec.lt_itinerary_link:
                url = rec.lt_itinerary_link.strip()
                if not (url.startswith("http://") or url.startswith("https://")):
                    url = "https://" + url
                rec.lt_itinerary_url = url
            else:
                rec.lt_itinerary_url = False

    # =========================
    # WRITE / CREATE / ACTIONS
    # =========================

    def write(self, vals):
        if "lt_itinerary_link" in vals:
            for rec in self:
                if rec.lt_itinerary_locked:
                    raise ValidationError(
                        _("Itinerary Link está BLOQUEADO. Desactiva el switch primero.")
                    )
        result = super().write(vals)
        for rec in self:
            rec._renumber_roster()
            rec._renumber_staff()
        return result

    @api.model
    def create(self, vals):
        code = vals.get("brigade_code") or "/"
        if code == "/":
            next_code = self.env["ir.sequence"].next_by_code("gb.brigade.code")
            if not next_code:
                raise ValidationError(
                    _(
                        "No se pudo obtener la secuencia 'gb.brigade.code'. "
                        "Verifica que 'sequence.xml' esté cargado en el manifest."
                    )
                )
            vals["brigade_code"] = next_code
        record = super().create(vals)
        record._renumber_roster()
        record._renumber_staff()
        return record

    def _renumber_roster(self):
        """Renumber all roster entries based on sequence."""
        self.ensure_one()
        roster_sorted = self.roster_ids.sorted(key=lambda r: (r.sequence, r.id))
        for idx, roster in enumerate(roster_sorted, start=1):
            if roster.line_number != idx:
                roster.write({'line_number': idx})

    def _renumber_staff(self):
        """Renumber all staff entries based on sequence."""
        self.ensure_one()
        staff_sorted = self.staff_ids.sorted(key=lambda s: (s.sequence, s.start_datetime or datetime.min, s.id))
        for idx, staff in enumerate(staff_sorted, start=1):
            if staff.line_number != idx:
                staff.write({'line_number': idx})

    def open_form_action(self):
        """Abre el formulario de la brigada desde la vista lista."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Brigade"),
            "res_model": "gb.brigade",
            "view_mode": "form",
            "res_id": self.id,
            "target": "current",
        }

    def action_view_roster_search(self):
        """
        Open Roster in a full window with search view enabled.
        """
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Roster - %s", self.name),
            "res_model": "gb.brigade.roster",
            "view_mode": "list,form",
            "domain": [("brigade_id", "=", self.id)],
            "context": dict(
                self.env.context,
                default_brigade_id=self.id,
                search_default_group_role=False,
            ),
            "target": "current",
        }

    def action_open_roster_import_wizard(self):
        """Abre el wizard de importación de Roster desde Excel."""
        self.ensure_one()
        action = self.env.ref(
            "global_brigades.action_gb_roster_import_wizard"
        ).read()[0]
        action["context"] = dict(
            self.env.context,
            default_brigade_id=self.id,
        )
        return action

    # =========================
    # S.A. NOTIFICATION ACTION
    # =========================
    def action_send_sa_notifications(self):
        """Send S.A. arrival notification emails to emergency contacts."""
        self.ensure_one()
        template = self.env.ref(
            'global_brigades.email_template_gb_sa_notification',
            raise_if_not_found=False,
        )
        if not template:
            raise UserError(
                _("S.A. email template not found. Please contact your administrator.")
            )

        pending = self.roster_ids.filtered(
            lambda r: r.sa and not r.sa_notified and r.partner_id.gb_emergency_contact_id
        )

        if not pending:
            raise UserError(
                _("No S.A. participants with pending notification and emergency contact found.")
            )

        sent = 0
        now = fields.Datetime.now()
        for roster in pending:
            emergency_contact = roster.partner_id.gb_emergency_contact_id
            if not emergency_contact.email:
                continue
            template.send_mail(
                roster.id,
                force_send=True,
                email_values={'email_to': emergency_contact.email},
            )
            roster.write({'sa_notified': now})
            sent += 1

        self.message_post(
            body=_("✉️ S.A. notifications sent to %s emergency contact(s).", sent),
            message_type='comment',
            subtype_xmlid='mail.mt_note',
        )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('S.A. Notifications Sent'),
                'message': _('%s email(s) sent to emergency contacts.', sent),
                'type': 'success',
                'sticky': False,
            },
        }

    # =========================
    # EXCEL EXPORTS
    # =========================

    def action_export_rooming_list(self):
        """
        Generate and download Rooming List Excel for this brigade.
        """
        self.ensure_one()

        booking_recs = self.env["gb.brigade.hotel.booking"].search(
            [("brigade_id", "=", self.id)],
            order="check_in_date, check_out_date, id"
        )

        if not booking_recs:
            raise UserError(_("No hotel bookings/rooming assignments found for this brigade. "
                            "Please create hotel bookings in the 'Hotels / Rooming' tab first."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Rooming List"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:M1")
        title_cell = ws["A1"]
        title_cell.value = f"ROOMING LIST - {self.name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "Check-In", "Check-Out", "Nights", "Hotel", "City",
            "Room #", "Room Type", "Beds", "Passenger Name", "Type",
            "Brigade Role", "Gender", "Passport",
        ]
        for col_num, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        row_idx = 4
        for booking in booking_recs:
            check_in_str = booking.check_in_date.strftime("%Y-%m-%d") if booking.check_in_date else ""
            check_out_str = booking.check_out_date.strftime("%Y-%m-%d") if booking.check_out_date else ""
            nights = booking.stay_nights or 0
            hotel_name = booking.partner_id.name if booking.partner_id else ""
            city = booking.city or ""

            all_passengers = []

            for line in booking.assignment_ids:
                for occupant in line.occupant_ids:
                    room_number = line.room_number or ""
                    room_type_val = dict(line.hotel_room_id._fields["room_type"].selection).get(
                        line.room_type, line.room_type or ""
                    ) if line.room_type else ""
                    bed_setup = line.bed_setup or ""
                    occupant_name = occupant.partner_id.name if occupant.partner_id else ""
                    brigade_role = occupant.brigade_role or ""
                    gender_val = ""
                    if occupant.gender:
                        gender_dict = dict(occupant.partner_id._fields["gb_gender"].selection)
                        gender_val = gender_dict.get(occupant.gender, occupant.gender)
                    passport = occupant.passport_no or ""
                    all_passengers.append({
                        "check_in": check_in_str, "check_out": check_out_str,
                        "nights": nights, "hotel": hotel_name, "city": city,
                        "room_number": room_number, "room_type": room_type_val,
                        "bed_setup": bed_setup, "name": occupant_name, "type": "Roster",
                        "brigade_role": brigade_role, "gender": gender_val, "passport": passport,
                    })

            for sline in booking.staff_assignment_ids:
                for staff_occupant in sline.occupant_staff_ids:
                    room_number = sline.room_number or ""
                    room_type_val = dict(sline.hotel_room_id._fields["room_type"].selection).get(
                        sline.room_type, sline.room_type or ""
                    ) if sline.room_type else ""
                    bed_setup = sline.bed_setup or ""
                    staff_name = staff_occupant.person_id.name if staff_occupant.person_id else ""
                    brigade_role = staff_occupant.brigade_role_default.name if staff_occupant.brigade_role_default else ""
                    gender_val = ""
                    if staff_occupant.gender:
                        gender_dict = dict(staff_occupant.person_id._fields["gb_gender"].selection)
                        gender_val = gender_dict.get(staff_occupant.gender, staff_occupant.gender)
                    passport = staff_occupant.person_id.gb_passport_no or ""
                    all_passengers.append({
                        "check_in": check_in_str, "check_out": check_out_str,
                        "nights": nights, "hotel": hotel_name, "city": city,
                        "room_number": room_number, "room_type": room_type_val,
                        "bed_setup": bed_setup, "name": staff_name, "type": "Staff",
                        "brigade_role": brigade_role, "gender": gender_val, "passport": passport,
                    })

            if not all_passengers:
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.border = border_thin
                    cell.alignment = center_alignment
                ws.cell(row=row_idx, column=1).value = check_in_str
                ws.cell(row=row_idx, column=2).value = check_out_str
                ws.cell(row=row_idx, column=3).value = nights
                ws.cell(row=row_idx, column=4).value = hotel_name
                ws.cell(row=row_idx, column=5).value = city
                row_idx += 1
            else:
                for pax in all_passengers:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_num)
                        cell.border = border_thin
                        cell.alignment = center_alignment if col_num <= 8 else Alignment(vertical="center")
                    ws.cell(row=row_idx, column=1).value = pax["check_in"]
                    ws.cell(row=row_idx, column=2).value = pax["check_out"]
                    ws.cell(row=row_idx, column=3).value = pax["nights"]
                    ws.cell(row=row_idx, column=4).value = pax["hotel"]
                    ws.cell(row=row_idx, column=5).value = pax["city"]
                    ws.cell(row=row_idx, column=6).value = pax["room_number"]
                    ws.cell(row=row_idx, column=7).value = pax["room_type"]
                    ws.cell(row=row_idx, column=8).value = pax["bed_setup"]
                    ws.cell(row=row_idx, column=9).value = pax["name"]
                    ws.cell(row=row_idx, column=10).value = pax["type"]
                    ws.cell(row=row_idx, column=11).value = pax["brigade_role"]
                    ws.cell(row=row_idx, column=12).value = pax["gender"]
                    ws.cell(row=row_idx, column=13).value = pax["passport"]
                    row_idx += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 25
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 20
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read())

        filename = f"Rooming_List_{self.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = self.env["ir.attachment"].create({
            "name": filename,
            "type": "binary",
            "datas": excel_data,
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_export_transport_list(self):
        """
        Generate and download Transport Assignments Excel for this brigade.
        """
        self.ensure_one()

        transport_recs = self.env["gb.brigade.transport"].search(
            [("brigade_id", "=", self.id)],
            order="date_time, id"
        )

        if not transport_recs:
            raise UserError(_("No transport records found for this brigade. "
                            "Please create transport records first."))

        wb = Workbook()
        ws = wb.active
        ws.title = "Transport Assignments"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = f"TRANSPORT ASSIGNMENTS - {self.name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "Date/Time", "Transport Title", "Origin", "Destination",
            "Vehicle", "Provider", "Capacity", "Passenger Name", "Type", "Role", "Notes"
        ]
        for col_num, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin

        row_idx = 4
        for transport in transport_recs:
            date_time = transport.date_time.strftime("%Y-%m-%d %H:%M") if transport.date_time else ""
            title = transport.title or ""
            origin = transport.origin or ""
            destination = transport.destination or ""
            transport_notes = transport.notes or ""

            if not transport.vehicle_line_ids:
                vehicle_name = transport.vehicle_id.name if transport.vehicle_id else ""
                provider_name = transport.provider_id.name if transport.provider_id else ""
                capacity = transport.vehicle_id.capacity if transport.vehicle_id else 0

                all_passengers = []
                for roster in transport.passenger_ids:
                    all_passengers.append({
                        'name': roster.partner_id.name if roster.partner_id else '',
                        'type': 'Roster',
                        'role': roster.brigade_role or ''
                    })
                for staff in transport.staff_passenger_ids:
                    all_passengers.append({
                        'name': staff.person_id.name if staff.person_id else '',
                        'type': 'Staff',
                        'role': dict(staff._fields['staff_role'].selection).get(staff.staff_role, staff.staff_role or '')
                    })

                if not all_passengers:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_num)
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical="center")
                    ws.cell(row=row_idx, column=1).value = date_time
                    ws.cell(row=row_idx, column=2).value = title
                    ws.cell(row=row_idx, column=3).value = origin
                    ws.cell(row=row_idx, column=4).value = destination
                    ws.cell(row=row_idx, column=5).value = vehicle_name
                    ws.cell(row=row_idx, column=6).value = provider_name
                    ws.cell(row=row_idx, column=7).value = capacity
                    ws.cell(row=row_idx, column=11).value = transport_notes
                    row_idx += 1
                else:
                    for pax in all_passengers:
                        for col_num in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_idx, column=col_num)
                            cell.border = border_thin
                            cell.alignment = Alignment(vertical="center")
                        ws.cell(row=row_idx, column=1).value = date_time
                        ws.cell(row=row_idx, column=2).value = title
                        ws.cell(row=row_idx, column=3).value = origin
                        ws.cell(row=row_idx, column=4).value = destination
                        ws.cell(row=row_idx, column=5).value = vehicle_name
                        ws.cell(row=row_idx, column=6).value = provider_name
                        ws.cell(row=row_idx, column=7).value = capacity
                        ws.cell(row=row_idx, column=8).value = pax['name']
                        ws.cell(row=row_idx, column=9).value = pax['type']
                        ws.cell(row=row_idx, column=10).value = pax['role']
                        ws.cell(row=row_idx, column=11).value = transport_notes
                        row_idx += 1
                continue

            for line in transport.vehicle_line_ids:
                vehicle_name = line.vehicle_id.name if line.vehicle_id else ""
                provider_name = line.provider_id.name if line.provider_id else ""
                capacity = line.capacity or 0

                vehicle_passengers = []
                for roster in line.roster_passenger_ids:
                    vehicle_passengers.append({
                        'name': roster.partner_id.name if roster.partner_id else '',
                        'type': 'Roster',
                        'role': roster.brigade_role or ''
                    })
                for staff in line.staff_passenger_ids:
                    vehicle_passengers.append({
                        'name': staff.person_id.name if staff.person_id else '',
                        'type': 'Staff',
                        'role': dict(staff._fields['staff_role'].selection).get(staff.staff_role, staff.staff_role or '')
                    })

                if not vehicle_passengers:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_num)
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical="center")
                    ws.cell(row=row_idx, column=1).value = date_time
                    ws.cell(row=row_idx, column=2).value = title
                    ws.cell(row=row_idx, column=3).value = origin
                    ws.cell(row=row_idx, column=4).value = destination
                    ws.cell(row=row_idx, column=5).value = vehicle_name
                    ws.cell(row=row_idx, column=6).value = provider_name
                    ws.cell(row=row_idx, column=7).value = capacity
                    ws.cell(row=row_idx, column=11).value = transport_notes
                    row_idx += 1
                else:
                    for pax in vehicle_passengers:
                        for col_num in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_idx, column=col_num)
                            cell.border = border_thin
                            cell.alignment = Alignment(vertical="center")
                        ws.cell(row=row_idx, column=1).value = date_time
                        ws.cell(row=row_idx, column=2).value = title
                        ws.cell(row=row_idx, column=3).value = origin
                        ws.cell(row=row_idx, column=4).value = destination
                        ws.cell(row=row_idx, column=5).value = vehicle_name
                        ws.cell(row=row_idx, column=6).value = provider_name
                        ws.cell(row=row_idx, column=7).value = capacity
                        ws.cell(row=row_idx, column=8).value = pax['name']
                        ws.cell(row=row_idx, column=9).value = pax['type']
                        ws.cell(row=row_idx, column=10).value = pax['role']
                        ws.cell(row=row_idx, column=11).value = transport_notes
                        row_idx += 1

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 25
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 20
        ws.column_dimensions["K"].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = base64.b64encode(output.read())

        filename = f"Transport_List_{self.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = self.env["ir.attachment"].create({
            "name": filename,
            "type": "binary",
            "datas": excel_data,
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
