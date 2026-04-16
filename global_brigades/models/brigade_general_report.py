# -*- coding: utf-8 -*-
# License LGPL-3.0 (https://www.gnu.org/licenses/lgpl-3.0.html)

import base64
from io import BytesIO
from datetime import datetime
import pytz
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise UserError(_("Please install openpyxl: pip install openpyxl"))


class GBBrigadeGeneralReport(models.Model):
    _name = 'gb.brigade.general.report'
    _description = 'Brigade General Report'
    _order = 'report_date desc, id desc'

    brigade_id = fields.Many2one(
        'gb.brigade',
        string='Brigade',
        required=True,
        ondelete='cascade',
    )

    report_date = fields.Datetime(
        string='Report Date',
        default=fields.Datetime.now,
        readonly=True,
    )

    excel_file = fields.Binary(
        string='Excel File',
        readonly=True,
    )

    excel_filename = fields.Char(
        string='Filename',
        readonly=True,
    )

    def _get_company_timezone(self):
        """Get the timezone of the company."""
        self.ensure_one()
        company = self.brigade_id.company_id or self.env.company
        if company.country_id and company.country_id.code == 'PA':
            return pytz.timezone('America/Panama')
        tz_name = self.env.user.tz or 'America/Panama'
        return pytz.timezone(tz_name)

    def _convert_to_local_time(self, utc_datetime):
        """Convert UTC datetime to company's local timezone."""
        if not utc_datetime:
            return None
        local_tz = self._get_company_timezone()
        if not utc_datetime.tzinfo:
            utc_datetime = pytz.utc.localize(utc_datetime)
        return utc_datetime.astimezone(local_tz)

    def _format_datetime(self, utc_datetime, fmt='%Y-%m-%d %H:%M'):
        """Format datetime in local timezone."""
        local_dt = self._convert_to_local_time(utc_datetime)
        return local_dt.strftime(fmt) if local_dt else ''

    def _format_date(self, date_value, fmt='%Y-%m-%d'):
        """Format date value."""
        return date_value.strftime(fmt) if date_value else ''

    def generate_excel_report(self):
        """Generate comprehensive Excel report for the brigade."""
        self.ensure_one()

        brigade = self.brigade_id
        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        self._create_brigade_info_sheet(wb, brigade)
        self._create_staff_sheet(wb, brigade)
        self._create_roster_sheet(wb, brigade)
        self._create_programs_activities_sheet(wb, brigade)
        self._create_arrivals_sheet(wb, brigade)
        self._create_departures_sheet(wb, brigade)
        self._create_hotels_sheet(wb, brigade)
        self._create_transport_sheet(wb, brigade)
        self._create_statistics_sheet(wb, brigade)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Brigade_Report_{brigade.brigade_code or brigade.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        self.write({
            'excel_file': base64.b64encode(output.read()),
            'excel_filename': filename,
        })

        return True

    # ------------------------------------------------------------------
    # SECTION STYLE HELPERS
    # ------------------------------------------------------------------

    def _section_title(self, ws, row, title, col_span=2):
        """Write a blue section title row."""
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFF', size=11)
        ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        if col_span > 1:
            ws.merge_cells(f'A{row}:{get_column_letter(col_span)}{row}')
        return row + 1

    def _kv_row(self, ws, row, label, value, label_width_col='A', value_col='B'):
        """Write a key-value row with light grey label."""
        ws[f'{label_width_col}{row}'] = label
        ws[f'{label_width_col}{row}'].font = Font(bold=True)
        ws[f'{label_width_col}{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws[f'{value_col}{row}'] = value
        return row + 1

    # ------------------------------------------------------------------
    # SHEET 1 — BRIGADE INFO  (mirrors PDF page 1)
    # ------------------------------------------------------------------

    def _create_brigade_info_sheet(self, wb, brigade):
        """Create Brigade Information sheet matching PDF page 1 layout."""
        ws = wb.create_sheet('Brigade Info')

        # ── Title ──
        ws['A1'] = f'{brigade.name}  —  Brigade Detailed Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        ws['A2'] = f'Generated: {datetime.now().strftime("%b %d, %Y")}'
        ws['A2'].font = Font(italic=True, size=10)

        row = 4

        # ── SUMMARY counts ──
        roster = brigade.roster_ids
        male_count = len(roster.filtered(lambda r: r.gender == 'male'))
        female_count = len(roster.filtered(lambda r: r.gender == 'female'))
        other_count = len(roster.filtered(lambda r: r.gender not in ('male', 'female') and r.gender))

        volunteer_only = len(roster)
        staff_total = len(brigade.staff_ids)
        total_participants = volunteer_only + staff_total

        row = self._section_title(ws, row, 'SUMMARY', 2)
        row = self._kv_row(ws, row, 'Volunteers', str(volunteer_only))
        row = self._kv_row(ws, row, 'Staff (Temporary)', str(staff_total))
        row = self._kv_row(ws, row, 'Total Participants', str(total_participants))
        row += 1

        # ── BRIGADE INFORMATION ──
        row = self._section_title(ws, row, 'BRIGADE INFORMATION', 2)
        row = self._kv_row(ws, row, 'Brigade Code', brigade.external_brigade_code or '')
        row = self._kv_row(ws, row, 'Internal Code', brigade.brigade_code or '')
        row = self._kv_row(ws, row, 'Arrival Date', self._format_date(brigade.arrival_date))
        row = self._kv_row(ws, row, 'Departure Date', self._format_date(brigade.departure_date))
        row = self._kv_row(ws, row, 'Program', brigade.brigade_program.name if brigade.brigade_program else '')
        row = self._kv_row(ws, row, 'Brigade Type', dict(brigade._fields['brigade_type'].selection).get(brigade.brigade_type, ''))
        row = self._kv_row(ws, row, 'Tier', dict(brigade._fields['brigade_tier'].selection).get(brigade.brigade_tier, '') if brigade.brigade_tier else '')
        row = self._kv_row(ws, row, 'Status', dict(brigade._fields['state'].selection).get(brigade.state, ''))
        row += 1

        # ── LOGISTICS AND COMPOUND ──
        row = self._section_title(ws, row, 'LOGISTICS AND COMPOUND', 2)
        row = self._kv_row(ws, row, 'Lodging Facility', brigade.lodging_facility_id.name if brigade.lodging_facility_id else '')
        communities = ', '.join(brigade.program_line_ids.mapped('community_id.name')) if brigade.program_line_ids else ''
        row = self._kv_row(ws, row, 'Communities', communities)
        row = self._kv_row(ws, row, 'Itinerary Link', brigade.lt_itinerary_link or '')
        row = self._kv_row(ws, row, 'Business Profile', brigade.business_profile_link or '')
        row += 1

        # ── GENDER BREAKDOWN ──
        row = self._section_title(ws, row, 'GENDER BREAKDOWN', 2)
        row = self._kv_row(ws, row, 'Male', str(male_count))
        row = self._kv_row(ws, row, 'Female', str(female_count))
        row = self._kv_row(ws, row, 'Other', str(other_count))
        row += 1

        # ── CONTACTS ──
        row = self._section_title(ws, row, 'CONTACTS', 2)
        program_advisors = ', '.join(brigade.program_associate_ids.mapped('name')) if brigade.program_associate_ids else ''
        row = self._kv_row(ws, row, 'Program Advisors', program_advisors)
        success_advisors = ', '.join(brigade.success_advisor_ids.mapped('name')) if brigade.success_advisor_ids else ''
        row = self._kv_row(ws, row, 'Success Advisor', success_advisors)
        sending_orgs = ', '.join(brigade.sending_organization_ids.mapped('name')) if brigade.sending_organization_ids else ''
        row = self._kv_row(ws, row, 'Sending Organization', sending_orgs)
        lead_coords = ', '.join(brigade.coordinator_ids.mapped('name')) if brigade.coordinator_ids else ''
        row = self._kv_row(ws, row, 'Lead Coordinator', lead_coords)
        chapter_presidents = ', '.join(brigade.chapter_president_faculty_ids.mapped('partner_id.name')) if brigade.chapter_president_faculty_ids else ''
        row = self._kv_row(ws, row, 'Chapter President(s)', chapter_presidents)
        professors = ', '.join(brigade.professor_chaperone_ids.mapped('partner_id.name')) if brigade.professor_chaperone_ids else ''
        row = self._kv_row(ws, row, 'Professor / Chaperone', professors)
        row += 1

        # ── ADDITIONAL INFORMATION ──
        row = self._section_title(ws, row, 'ADDITIONAL INFORMATION & OBSERVATIONS', 2)
        ws[f'A{row}'] = brigade.extra_info or ''
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:B{row}')
        ws.row_dimensions[row].height = 60

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 60

    # ------------------------------------------------------------------
    # SHEET 2 — STAFF ROSTER  (mirrors PDF page 2)
    # ------------------------------------------------------------------

    def _create_staff_sheet(self, wb, brigade):
        """Create Staff sheet matching PDF page 2 layout."""
        ws = wb.create_sheet('Staff')

        # Title
        ws['A1'] = f'TEMPORARY STAFF ROSTER  ({len(brigade.staff_ids)} MEMBERS)'
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:G1')

        headers = ['#', 'Role', 'Name', 'Diet / Restrictions', 'Allergy', 'Start', 'End']
        self._write_sheet_headers(ws, headers, start_row=2)

        row = 3
        for staff in brigade.staff_ids.sorted(key=lambda s: (s.sequence, s.id)):
            brigade_role = ''
            if staff.brigade_role_default:
                try:
                    role_dict = dict(staff.person_id._fields['gb_brigade_role'].selection)
                    brigade_role = role_dict.get(staff.brigade_role_default, str(staff.brigade_role_default))
                except Exception:
                    brigade_role = str(staff.brigade_role_default)

            ws[f'A{row}'] = staff.line_number or (row - 2)
            ws[f'B{row}'] = brigade_role
            ws[f'C{row}'] = staff.person_id.name or ''
            ws[f'D{row}'] = staff.diet or ''
            ws[f'E{row}'] = staff.allergy or ''
            ws[f'F{row}'] = self._format_date(staff.start_datetime.date() if staff.start_datetime else None)
            ws[f'G{row}'] = self._format_date(staff.end_datetime.date() if staff.end_datetime else None)
            row += 1

        col_widths = [5, 22, 28, 30, 28, 12, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # SHEET 3 — VOLUNTEER ROSTER  (mirrors PDF page 3)
    # ------------------------------------------------------------------

    def _create_roster_sheet(self, wb, brigade):
        """Create Volunteer Roster sheet matching PDF page 3 layout."""
        ws = wb.create_sheet('Volunteers')

        roster_list = brigade.roster_ids.sorted(key=lambda r: (r.sequence, r.id))

        ws['A1'] = f'VOLUNTEER ROSTER  ({len(roster_list)} PARTICIPANTS)'
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:L1')

        headers = [
            '#', 'Name', 'Gender', 'DOB', 'Role',
            'Passport #', 'Exp.', 'Spanish',
            'Diet', 'Medical / Allergy', 'MED', 'T-Shirt'
        ]
        self._write_sheet_headers(ws, headers, start_row=2)

        row = 3
        for roster in roster_list:
            gender_map = {'male': 'M', 'female': 'F', 'other': 'Other'}
            gender_val = gender_map.get(roster.gender, '') if roster.gender else ''

            # Combine medical condition + allergy in one column
            med_allergy_parts = []
            if roster.medical_condition:
                med_allergy_parts.append(roster.medical_condition)
            if roster.allergy:
                med_allergy_parts.append(roster.allergy)
            med_allergy = ' | '.join(med_allergy_parts)

            ws[f'A{row}'] = roster.line_number or (row - 2)
            ws[f'B{row}'] = roster.partner_id.name or ''
            ws[f'C{row}'] = gender_val
            ws[f'D{row}'] = self._format_date(roster.birthdate)
            ws[f'E{row}'] = roster.brigade_role or ''
            ws[f'F{row}'] = roster.passport_no or ''
            ws[f'G{row}'] = self._format_date(roster.passport_expiry)
            ws[f'H{row}'] = 'Yes' if roster.spanish_speaker else 'No'
            ws[f'I{row}'] = roster.diet or ''
            ws[f'J{row}'] = med_allergy
            ws[f'K{row}'] = roster.medications or ''
            ws[f'L{row}'] = dict(roster.partner_id._fields['gb_tshirt_size'].selection).get(roster.tshirt_size, '') if roster.tshirt_size else ''
            row += 1

        col_widths = [5, 28, 8, 12, 18, 14, 12, 9, 25, 35, 30, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # SHEET 4 — PROGRAMS & ACTIVITIES  (unified table with Type column)
    # ------------------------------------------------------------------

    def _create_programs_activities_sheet(self, wb, brigade):
        """Create unified Programs & Activities sheet matching PDF table."""
        ws = wb.create_sheet('Programs & Activities')

        ws['A1'] = 'PROGRAM & ACTIVITIES'
        ws['A1'].font = Font(bold=True, size=13)
        ws.merge_cells('A1:H1')

        headers = ['Program / Activity', 'Type', 'Start', 'End', 'Place', 'Responsible', '#', 'Notes']
        self._write_sheet_headers(ws, headers, start_row=2)

        row = 3

        # Programs from program_line_ids
        for program in brigade.program_line_ids:
            program_type = program.program_id.name if program.program_id else 'Program'
            ws[f'A{row}'] = program.program_id.name if program.program_id else ''
            ws[f'B{row}'] = program_type
            ws[f'C{row}'] = self._format_date(program.start_date)
            ws[f'D{row}'] = self._format_date(program.end_date)
            ws[f'E{row}'] = program.community_id.name if program.community_id else (program.location or '')
            ws[f'F{row}'] = program.coordinator_id.name if program.coordinator_id else ''
            ws[f'G{row}'] = ''
            ws[f'H{row}'] = program.notes or ''
            row += 1

        # Activities from brigade_activity_ids
        for activity in brigade.brigade_activity_ids:
            tags = ', '.join(activity.tag_ids.mapped('name')) if activity.tag_ids else ''
            ws[f'A{row}'] = activity.name or ''
            ws[f'B{row}'] = tags or 'Activity'
            ws[f'C{row}'] = self._format_datetime(activity.start_datetime, fmt='%Y-%m-%d')
            ws[f'D{row}'] = self._format_datetime(activity.end_datetime, fmt='%Y-%m-%d')
            ws[f'E{row}'] = activity.place or ''
            ws[f'F{row}'] = activity.responsible_id.name if activity.responsible_id else ''
            ws[f'G{row}'] = activity.participant_count
            ws[f'H{row}'] = activity.notes or ''
            row += 1

        col_widths = [30, 18, 12, 12, 22, 22, 6, 35]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # SHEET 5 — ARRIVALS
    # ------------------------------------------------------------------

    def _create_arrivals_sheet(self, wb, brigade):
        """Create Arrivals sheet."""
        ws = wb.create_sheet('Arrivals')

        headers = [
            'Title', 'Flight Number', 'Arrival DateTime', 'Through SAP',
            'Arrival Hotel', 'Hotel City/Time', '# Pax', 'Passengers',
            'Special Transport', 'Extra Charge'
        ]
        self._write_sheet_headers(ws, headers)

        row = 2
        for arrival in brigade.arrival_ids:
            ws[f'A{row}'] = arrival.title or ''
            ws[f'B{row}'] = arrival.flight_number or ''
            ws[f'C{row}'] = self._format_datetime(arrival.date_time_arrival)
            ws[f'D{row}'] = arrival.flight_through_sap or ''
            ws[f'E{row}'] = arrival.arrival_hotel_id.name if arrival.arrival_hotel_id else ''
            ws[f'F{row}'] = arrival.arrival_hotel_city_time or ''
            ws[f'G{row}'] = arrival.n_pax
            ws[f'H{row}'] = ', '.join(arrival.passenger_ids.mapped('partner_id.name'))
            ws[f'I{row}'] = 'Yes' if arrival.special_transport else 'No'
            ws[f'J{row}'] = arrival.extra_charge or ''
            row += 1

        self._auto_adjust_columns(ws)

    # ------------------------------------------------------------------
    # SHEET 6 — DEPARTURES
    # ------------------------------------------------------------------

    def _create_departures_sheet(self, wb, brigade):
        """Create Departures sheet."""
        ws = wb.create_sheet('Departures')

        headers = [
            'Title', 'Flight Number', 'Departure DateTime', 'Through SAP',
            'Departure Hotel', 'Hotel City', '# Pax', 'Passengers',
            'Special Transport', 'Extra Charge'
        ]
        self._write_sheet_headers(ws, headers)

        row = 2
        for departure in brigade.departure_ids:
            ws[f'A{row}'] = departure.title or ''
            ws[f'B{row}'] = departure.flight_number or ''
            ws[f'C{row}'] = self._format_datetime(departure.date_time_departure)
            ws[f'D{row}'] = departure.flight_through_sap or ''
            ws[f'E{row}'] = departure.departure_hotel_id.name if departure.departure_hotel_id else ''
            ws[f'F{row}'] = departure.departure_hotel_city or ''
            ws[f'G{row}'] = departure.n_pax
            ws[f'H{row}'] = ', '.join(departure.passenger_ids.mapped('partner_id.name'))
            ws[f'I{row}'] = 'Yes' if departure.special_transport else 'No'
            ws[f'J{row}'] = departure.extra_charge or ''
            row += 1

        self._auto_adjust_columns(ws)

    # ------------------------------------------------------------------
    # SHEET 7 — HOTELS
    # ------------------------------------------------------------------

    def _create_hotels_sheet(self, wb, brigade):
        """Create Hotels sheet."""
        ws = wb.create_sheet('Hotels')

        headers = [
            'Hotel', 'Check-in Date', 'Check-out Date', 'Stay Nights',
            '# Rooms', 'Room Distribution', 'Total Guests', 'Notes'
        ]
        self._write_sheet_headers(ws, headers)

        row = 2
        for booking in brigade.hotel_booking_ids:
            total_rooms = len(booking.assignment_ids) + len(booking.staff_assignment_ids)
            room_distribution = []
            for line in booking.assignment_ids:
                room_distribution.append(f"{line.room_number or 'N/A'} ({line.pax_count} pax)")
            for sline in booking.staff_assignment_ids:
                room_distribution.append(f"{sline.room_number or 'N/A'} ({sline.pax_count} staff)")

            ws[f'A{row}'] = booking.partner_id.name if booking.partner_id else ''
            ws[f'B{row}'] = self._format_date(booking.check_in_date)
            ws[f'C{row}'] = self._format_date(booking.check_out_date)
            ws[f'D{row}'] = int(booking.stay_nights)
            ws[f'E{row}'] = total_rooms
            ws[f'F{row}'] = ', '.join(room_distribution) if room_distribution else ''
            ws[f'G{row}'] = booking.total_headcount
            ws[f'H{row}'] = booking.note or ''
            row += 1

        self._auto_adjust_columns(ws)

    # ------------------------------------------------------------------
    # SHEET 8 — TRANSPORT
    # ------------------------------------------------------------------

    def _create_transport_sheet(self, wb, brigade):
        """Create Transport sheet."""
        ws = wb.create_sheet('Transport')

        headers = [
            'Transport Title', 'Date/Time', 'Provider', 'Vehicle',
            'Origin', 'Destination', '# Vehicles', '# Passengers', 'Notes'
        ]
        self._write_sheet_headers(ws, headers)

        row = 2
        for transport in brigade.transport_ids:
            ws[f'A{row}'] = transport.title or ''
            ws[f'B{row}'] = self._format_datetime(transport.date_time)
            ws[f'C{row}'] = transport.provider_id.name if transport.provider_id else ''
            ws[f'D{row}'] = transport.vehicle_id.name if transport.vehicle_id else ''
            ws[f'E{row}'] = transport.origin or ''
            ws[f'F{row}'] = transport.destination or ''
            ws[f'G{row}'] = transport.vehicle_count
            ws[f'H{row}'] = transport.n_pax
            ws[f'I{row}'] = transport.notes or ''
            row += 1

        self._auto_adjust_columns(ws)

    # ------------------------------------------------------------------
    # SHEET 9 — STATISTICS
    # ------------------------------------------------------------------

    def _create_statistics_sheet(self, wb, brigade):
        """Create Statistics sheet with gender breakdown."""
        ws = wb.create_sheet('Statistics')

        ws['A1'] = 'BRIGADE STATISTICS'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        row = 3

        roster = brigade.roster_ids
        male_count = len(roster.filtered(lambda r: r.gender == 'male'))
        female_count = len(roster.filtered(lambda r: r.gender == 'female'))
        other_count = len(roster.filtered(lambda r: r.gender not in ('male', 'female') and r.gender))
        spanish_count = len(roster.filtered(lambda r: r.spanish_speaker))

        communities = ', '.join(
            list(dict.fromkeys(brigade.program_line_ids.mapped('community_id.name')))
        ) if brigade.program_line_ids else ''

        stats_data = [
            ('— PARTICIPANTS —', ''),
            ('Total Volunteers', str(len(roster))),
            ('Total Staff', str(brigade.staff_count)),
            ('Total Participants', str(brigade.total_participants)),
            ('— GENDER BREAKDOWN —', ''),
            ('Male', str(male_count)),
            ('Female', str(female_count)),
            ('Other', str(other_count)),
            ('— VOLUNTEERS INFO —', ''),
            ('Spanish Speakers', str(spanish_count)),
            ('— LOGISTICS —', ''),
            ('Communities', communities),
            ('# Arrivals', str(brigade.arrival_count)),
            ('# Departures', str(brigade.departure_count)),
            ('# Hotel Blocks', str(brigade.hotel_booking_count)),
            ('Total Hotel Nights', str(brigade.total_stay_nights)),
            ('— STAFF BREAKDOWN —', ''),
            ('Medical Staff', str(brigade.medical_staff_count)),
            ('Dental Staff', str(brigade.dental_staff_count)),
            ('Logistics Staff', str(brigade.logistics_staff_count)),
            ('Translators / Interpreters', str(brigade.translator_staff_count)),
        ]

        section_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        section_font = Font(bold=True, color='FFFFFF')

        for label, value in stats_data:
            if label.startswith('—') and label.endswith('—'):
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = section_font
                ws[f'A{row}'].fill = section_fill
                ws.merge_cells(f'A{row}:B{row}')
            else:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                ws[f'B{row}'] = value
            row += 1

        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 50

    # ------------------------------------------------------------------
    # SHARED HELPERS
    # ------------------------------------------------------------------

    def _write_sheet_headers(self, ws, headers, start_row=1):
        """Write headers to worksheet with styling."""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
