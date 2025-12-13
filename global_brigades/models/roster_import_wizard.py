# -*- coding: utf-8 -*-

# License LGPL-3.0 (https://www.gnu.org/licenses/lgpl-3.0.html)

import base64
import io
import openpyxl
from openpyxl.utils import get_column_letter
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError


class LTRosterImportWizard(models.TransientModel):
    """
    Wizard to import roster members from Excel file.
    Creates Partner contacts automatically and adds them to brigade roster.
    """
    _name = "lt.roster.import.wizard"
    _description = "Import Roster from Excel"

    brigade_id = fields.Many2one(
        "gb.brigade",
        string="Brigade",
        required=True,
        help="Brigade to import roster into.",
    )

    excel_file = fields.Binary(
        string="Excel File",
        required=True,
        help="Upload Excel file with roster data (use template provided).",
    )

    filename = fields.Char(
        string="Filename",
        help="Name of the uploaded file.",
    )

    import_count = fields.Integer(
        string="Records to Import",
        readonly=True,
        help="Number of valid records found in the file.",
    )

    skip_duplicates = fields.Boolean(
        string="Skip Duplicates",
        default=True,
        help="If True, skip records where partner already exists (by email).",
    )

    create_partners = fields.Boolean(
        string="Create Partners",
        default=True,
        help="If True, create new partner contacts for people not in system.",
    )

    @api.onchange("excel_file", "filename")
    def _onchange_excel_file(self):
        """
        When file is uploaded, check and count valid rows.
        """
        if not self.excel_file:
            self.import_count = 0
            return

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data))

            # Get the first sheet (Roster Import)
            if "Roster Import" in wb.sheetnames:
                ws = wb["Roster Import"]
            else:
                ws = wb.active

            # Count rows (skip header)
            row_count = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                # Check if row has data (at least Name field should be filled)
                if row[0] and row[0].value:
                    row_count += 1

            self.import_count = row_count

        except Exception as e:
            raise ValidationError(
                _("Error reading Excel file: %s. Make sure it's a valid .xlsx file.") % str(e)
            )

    def action_import_roster(self):
        """
        Main import action: read Excel and create Partners + Roster records.
        """
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_("Please select an Excel file to import."))

        if not self.brigade_id:
            raise UserError(_("Please select a brigade."))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data))

            # Get the first sheet
            if "Roster Import" in wb.sheetnames:
                ws = wb["Roster Import"]
            else:
                ws = wb.active

            # Read headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value.lower().replace(" ", "_") if cell.value else None)

            # Mapping of Excel column names to res.partner fields
            partner_field_mapping = {
                "name": "name",
                "email": "email",
                "phone": "phone",
                "mobile": "mobile",
                "gender": "gb_gender",
                "birth_date": "gb_birthdate",
                "spanish_speaker": "gb_spanish_speaker",
                "passport_number": "gb_passport_no",
                "passport_expiry": "gb_passport_expiry",
                "citizenship": "gb_citizenship",
                "diet_restrictions": "gb_diet",
                "medical_condition": "gb_medical_condition",
                "medications": "gb_medications",
                "allergies": "gb_allergy",
                "t-shirt_size": "gb_tshirt_size",
            )

            # Mapping for emergency contact
            emergency_contact_mapping = {
                "emergency_contact_name": "emergency_contact_name",
                "emergency_contact_email": "emergency_contact_email",
            }

            created_count = 0
            updated_count = 0
            skipped_count = 0
            errors = []

            # Process each row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Extract row data
                    row_data = {}
                    for col_num, cell in enumerate(row):
                        if col_num < len(headers) and headers[col_num]:
                            row_data[headers[col_num]] = cell.value

                    # Skip empty rows
                    if not row_data.get("name"):
                        continue

                    # Prepare partner values
                    partner_vals = {"is_company": False}

                    for excel_field, odoo_field in partner_field_mapping.items():
                        if excel_field in row_data and row_data[excel_field]:
                            value = row_data[excel_field]

                            # Handle boolean fields
                            if odoo_field in ["gb_spanish_speaker"]:
                                if isinstance(value, str):
                                    value = value.lower() in ["true", "1", "yes"]
                                elif isinstance(value, bool):
                                    pass
                                else:
                                    value = False

                            partner_vals[odoo_field] = value

                    # Handle emergency contact (create or link)
                    emergency_contact_id = False
                    if row_data.get("emergency_contact_name"):
                        # Search or create emergency contact
                        emergency_email = row_data.get("emergency_contact_email")
                        emerg_search = self.env["res.partner"].search(
                            [("name", "=", row_data["emergency_contact_name"]),
                             ("email", "=", emergency_email)] if emergency_email
                            else [("name", "=", row_data["emergency_contact_name"])]
                        )

                        if emerg_search:
                            emergency_contact_id = emerg_search[0].id
                        else:
                            emerg_vals = {"name": row_data["emergency_contact_name"],
                                          "is_company": False}
                            if emergency_email:
                                emerg_vals["email"] = emergency_email

                            emerg_partner = self.env["res.partner"].create(emerg_vals)
                            emergency_contact_id = emerg_partner.id

                    # Check if partner already exists (by email or name)
                    email = partner_vals.get("email")
                    name = partner_vals.get("name")

                    if email:
                        partner = self.env["res.partner"].search([("email", "=", email)], limit=1)
                    else:
                        partner = self.env["res.partner"].search([("name", "=", name)], limit=1)

                    if partner:
                        if self.skip_duplicates:
                            skipped_count += 1
                            continue
                        else:
                            # Update existing partner
                            partner.write(partner_vals)
                            updated_count += 1
                            partner_id = partner.id
                    else:
                        if not self.create_partners:
                            skipped_count += 1
                            continue

                        # Create new partner
                        partner = self.env["res.partner"].create(partner_vals)
                        created_count += 1
                        partner_id = partner.id

                    # Create roster record linked to this brigade
                    roster_vals = {
                        "brigade_id": self.brigade_id.id,
                        "partner_id": partner_id,
                        "brigade_role": row_data.get("brigade_role", ""),
                        "sa": row_data.get("s.a.", False),
                    }

                    if row_data.get("s.a."):
                        if isinstance(row_data["s.a."], str):
                            roster_vals["sa"] = row_data["s.a."].lower() in ["true", "1", "yes"]

                    # Check if roster record already exists for this brigade+partner
                    existing_roster = self.env["gb.brigade.roster"].search(
                        [("brigade_id", "=", self.brigade_id.id),
                         ("partner_id", "=", partner_id)]
                    )

                    if not existing_roster:
                        self.env["gb.brigade.roster"].create(roster_vals)

                except Exception as e:
                    errors.append(_("Row %d: %s") % (row_num, str(e)))
                    continue

            # Prepare result message
            message = _("Import completed!\n\n")
            message += _("Created Partners: %d\n") % created_count
            message += _("Updated Partners: %d\n") % updated_count
            message += _("Skipped: %d\n") % skipped_count

            if errors:
                message += _("\nErrors:\n")
                for error in errors[:10]:  # Show first 10 errors
                    message += error + "\n"
                if len(errors) > 10:
                    message += _("... and %d more errors") % (len(errors) - 10)

            raise UserError(message)

        except UserError:
            raise
        except Exception as e:
            raise UserError(
                _("Error processing Excel file: %s") % str(e)
            )
