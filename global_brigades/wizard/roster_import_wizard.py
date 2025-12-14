# -*- coding: utf-8 -*-
import base64
from io import BytesIO
from datetime import datetime, date

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class GBRosterImportWizard(models.TransientModel):
    _name = "gb.roster.import.wizard"
    _description = "Import Roster from Excel"

    brigade_id = fields.Many2one(
        "gb.brigade",
        string="Brigade",
        required=True,
        ondelete="cascade",
    )

    # IMPORTANT: NOT required here, so "Download Template" works without uploading a file.
    upload_file = fields.Binary(string="Excel File (.xlsx)", required=False)
    filename = fields.Char(string="Filename")

    create_missing_partners = fields.Boolean(
        string="Create contact if email not found",
        default=True,
    )
    update_existing_partners = fields.Boolean(
        string="Update contact data if email exists",
        default=False,
        help="If checked, updates partner fields from the Excel row (basic data and GB profile fields).",
    )

    # -------------------------
    # Helpers
    # -------------------------
    def _require_openpyxl(self):
        try:
            import openpyxl  # noqa: F401
        except Exception:
            raise UserError(_(
                "Missing python dependency: openpyxl.\n\n"
                "Install it on the server, for example:\n"
                "  sudo -u odoo -H /opt/odoo/venv/bin/pip install openpyxl\n"
            ))

    def _normalize_email(self, email):
        email = (email or "").strip()
        return email.lower()

    def _parse_bool(self, v):
        if v is None:
            return False
        if isinstance(v, bool):
            return v
        sval = str(v).strip().lower()
        return sval in ("1", "true", "yes", "y", "si", "sí", "x")

    def _parse_date(self, v):
        if not v:
            return False
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.date()
        sval = str(v).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(sval, fmt).date()
            except Exception:
                continue
        raise UserError(_("Invalid date value: %s (use YYYY-MM-DD or DD/MM/YYYY)") % sval)

    def _map_gender(self, v):
        if not v:
            return False
        sval = str(v).strip().lower()
        mapping = {
            "male": "male",
            "m": "male",
            "hombre": "male",
            "masculino": "male",
            "female": "female",
            "f": "female",
            "mujer": "female",
            "femenino": "female",
            "other": "other",
            "otro": "other",
            "otra": "other",
        }
        return mapping.get(sval, sval)

    def _map_tshirt(self, v):
        if not v:
            return False
        sval = str(v).strip().lower()
        mapping = {
            "xs": "xs",
            "s": "s",
            "m": "m",
            "l": "l",
            "xl": "xl",
            "xxl": "xxl",
            "2xl": "xxl",
        }
        return mapping.get(sval, sval)

    # -------------------------
    # Template download
    # -------------------------
    def _build_template_xlsx(self):
        """Return xlsx bytes for the roster import template."""
        self._require_openpyxl()
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"

        headers = [
            "email",
            "name",
            "phone",
            "mobile",
            "gender",
            "birthdate",
            "spanish_speaker",
            "passport_no",
            "passport_expiry",
            "citizenship",
            "tshirt_size",
            "brigade_role",
            "sa",
            "diet",
            "medical_condition",
            "medications",
            "allergy",
            "emergency_contact_email",
            "emergency_contact_name",
            "emergency_contact_phone",
            "emergency_contact_mobile",
        ]
        ws.append(headers)

        # Example row (helps users)
        ws.append([
            "john.doe@email.com",
            "John Doe",
            "+591 70000000",
            "+591 70000001",
            "male",
            "1990-01-31",
            "yes",
            "P1234567",
            "2030-12-31",
            "Bolivia",
            "m",
            "Volunteer",
            "no",
            "None",
            "None",
            "None",
            "None",
            "jane.doe@email.com",
            "Jane Doe",
            "+591 70000002",
            "+591 70000003",
        ])

        bold = Font(bold=True)
        for col_idx, _h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def action_download_template(self):
        """Download the Excel template from the wizard."""
        self.ensure_one()

        xlsx_bytes = self._build_template_xlsx()
        b64 = base64.b64encode(xlsx_bytes)

        attachment = self.env["ir.attachment"].sudo().create({
            "name": "gb_roster_import_template.xlsx",
            "type": "binary",
            "datas": b64,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": self._name,
            "res_id": self.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }

    # -------------------------
    # Import
    # -------------------------
    def action_import(self):
        self.ensure_one()
        self._require_openpyxl()

        if not self.upload_file:
            raise UserError(_("Please upload an Excel file before importing."))

        import openpyxl

        content = base64.b64decode(self.upload_file)
        try:
            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as e:
            raise UserError(_("Could not read the Excel file. Error: %s") % (e,))

        ws = wb.active

        # Read header
        header_row = []
        for cell in ws[1]:
            header_row.append((str(cell.value).strip() if cell.value is not None else "").strip())

        if not any(header_row):
            raise UserError(_("The Excel file seems empty (no header row found)."))

        def norm(h):
            return (h or "").strip().lower()

        col_map = {norm(h): idx for idx, h in enumerate(header_row)}

        # Required: email
        if "email" not in col_map:
            raise UserError(_(
                "Missing required column: 'email'.\n\n"
                "Minimum header:\n"
                "  email\n\n"
                "Recommended full header:\n"
                "  email, name, phone, mobile, gender, birthdate, spanish_speaker,\n"
                "  passport_no, passport_expiry, citizenship, tshirt_size,\n"
                "  brigade_role, sa, diet, medical_condition, medications, allergy,\n"
                "  emergency_contact_email, emergency_contact_name, emergency_contact_phone, emergency_contact_mobile\n"
            ))

        # Optional (contact basics)
        name_col = col_map.get("name")
        phone_col = col_map.get("phone")
        mobile_col = col_map.get("mobile")

        # Roster-only
        brigade_role_col = col_map.get("brigade_role")
        sa_col = col_map.get("sa")

        # GB profile fields (stored on res.partner)
        gender_col = col_map.get("gender")
        birthdate_col = col_map.get("birthdate")
        spanish_speaker_col = col_map.get("spanish_speaker")
        passport_no_col = col_map.get("passport_no")
        passport_expiry_col = col_map.get("passport_expiry")
        citizenship_col = col_map.get("citizenship")
        tshirt_size_col = col_map.get("tshirt_size")
        diet_col = col_map.get("diet")
        medical_condition_col = col_map.get("medical_condition")
        medications_col = col_map.get("medications")
        allergy_col = col_map.get("allergy")

        # Emergency contact (res.partner)
        emergency_contact_email_col = col_map.get("emergency_contact_email")
        emergency_contact_name_col = col_map.get("emergency_contact_name")
        emergency_contact_phone_col = col_map.get("emergency_contact_phone")
        emergency_contact_mobile_col = col_map.get("emergency_contact_mobile")

        Partner = self.env["res.partner"].sudo()
        Roster = self.env["gb.brigade.roster"].sudo()

        created_partners = 0
        updated_partners = 0
        created_emergency_contacts = 0
        created_roster = 0
        skipped_existing = 0

        errors = []

        def get_cell(row, idx):
            if idx is None:
                return None
            if idx >= len(row):
                return None
            return row[idx].value

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            email = self._normalize_email(get_cell(row, col_map["email"]))

            # Skip blank lines
            if not email:
                if all((c.value is None or str(c.value).strip() == "") for c in row):
                    continue
                errors.append(_("Row %s: missing email") % row_idx)
                continue

            # Basic values
            name = (str(get_cell(row, name_col)).strip() if get_cell(row, name_col) is not None else "").strip()
            phone = (str(get_cell(row, phone_col)).strip() if get_cell(row, phone_col) is not None else "").strip()
            mobile = (str(get_cell(row, mobile_col)).strip() if get_cell(row, mobile_col) is not None else "").strip()

            # Roster values
            brigade_role = (str(get_cell(row, brigade_role_col)).strip() if get_cell(row, brigade_role_col) is not None else "").strip()
            sa = self._parse_bool(get_cell(row, sa_col))

            # GB profile values (partner)
            gender = self._map_gender(get_cell(row, gender_col))
            birthdate = self._parse_date(get_cell(row, birthdate_col)) if birthdate_col is not None else False
            spanish_speaker = self._parse_bool(get_cell(row, spanish_speaker_col)) if spanish_speaker_col is not None else False
            passport_no = (str(get_cell(row, passport_no_col)).strip() if get_cell(row, passport_no_col) is not None else "").strip()
            passport_expiry = self._parse_date(get_cell(row, passport_expiry_col)) if passport_expiry_col is not None else False
            citizenship = (str(get_cell(row, citizenship_col)).strip() if get_cell(row, citizenship_col) is not None else "").strip()
            tshirt_size = self._map_tshirt(get_cell(row, tshirt_size_col))
            diet = (str(get_cell(row, diet_col)).strip() if get_cell(row, diet_col) is not None else "").strip()
            medical_condition = (str(get_cell(row, medical_condition_col)).strip() if get_cell(row, medical_condition_col) is not None else "").strip()
            medications = (str(get_cell(row, medications_col)).strip() if get_cell(row, medications_col) is not None else "").strip()
            allergy = (str(get_cell(row, allergy_col)).strip() if get_cell(row, allergy_col) is not None else "").strip()

            # Emergency contact values
            ec_email = self._normalize_email(get_cell(row, emergency_contact_email_col)) if emergency_contact_email_col is not None else ""
            ec_name = (str(get_cell(row, emergency_contact_name_col)).strip() if get_cell(row, emergency_contact_name_col) is not None else "").strip()
            ec_phone = (str(get_cell(row, emergency_contact_phone_col)).strip() if get_cell(row, emergency_contact_phone_col) is not None else "").strip()
            ec_mobile = (str(get_cell(row, emergency_contact_mobile_col)).strip() if get_cell(row, emergency_contact_mobile_col) is not None else "").strip()

            # Find/create main partner by email
            partner = Partner.search([("email", "=", email)], limit=1)
            if not partner:
                if not self.create_missing_partners:
                    errors.append(_("Row %s: email '%s' not found in contacts and 'Create contact' is disabled")
                                  % (row_idx, email))
                    continue

                vals = {"name": name or email, "email": email}
                if phone:
                    vals["phone"] = phone
                if mobile:
                    vals["mobile"] = mobile
                partner = Partner.create(vals)
                created_partners += 1

            # Create/find emergency contact if provided
            emergency_contact = False
            if ec_email:
                emergency_contact = Partner.search([("email", "=", ec_email)], limit=1)
                if not emergency_contact:
                    emergency_contact_vals = {"name": ec_name or ec_email, "email": ec_email}
                    if ec_phone:
                        emergency_contact_vals["phone"] = ec_phone
                    if ec_mobile:
                        emergency_contact_vals["mobile"] = ec_mobile
                    emergency_contact = Partner.create(emergency_contact_vals)
                    created_emergency_contacts += 1

            # Update partner data if enabled
            if self.update_existing_partners:
                upd = {}

                # Basics (only set if provided)
                if name and (partner.name or "").strip() != name:
                    upd["name"] = name
                if phone and (partner.phone or "").strip() != phone:
                    upd["phone"] = phone
                if mobile and (partner.mobile or "").strip() != mobile:
                    upd["mobile"] = mobile

                # GB profile fields (only set if provided)
                if gender:
                    upd["gb_gender"] = gender
                if birthdate:
                    upd["gb_birthdate"] = birthdate
                if spanish_speaker_col is not None:
                    upd["gb_spanish_speaker"] = spanish_speaker
                if passport_no:
                    upd["gb_passport_no"] = passport_no
                if passport_expiry:
                    upd["gb_passport_expiry"] = passport_expiry
                if citizenship:
                    upd["gb_citizenship"] = citizenship
                if tshirt_size:
                    upd["gb_tshirt_size"] = tshirt_size
                if diet:
                    upd["gb_diet"] = diet
                if medical_condition:
                    upd["gb_medical_condition"] = medical_condition
                if medications:
                    upd["gb_medications"] = medications
                if allergy:
                    upd["gb_allergy"] = allergy
                if emergency_contact:
                    upd["gb_emergency_contact_id"] = emergency_contact.id

                if upd:
                    partner.write(upd)
                    updated_partners += 1
            else:
                # Light-touch: set emergency contact only if empty and provided
                if emergency_contact and not partner.gb_emergency_contact_id:
                    partner.write({"gb_emergency_contact_id": emergency_contact.id})

            # Avoid duplicate roster line for same brigade+partner
            existing = Roster.search([
                ("brigade_id", "=", self.brigade_id.id),
                ("partner_id", "=", partner.id),
            ], limit=1)

            if existing:
                skipped_existing += 1
                continue

            roster_vals = {
                "brigade_id": self.brigade_id.id,
                "partner_id": partner.id,
                "sa": sa,
            }
            if brigade_role:
                roster_vals["brigade_role"] = brigade_role

            Roster.create(roster_vals)
            created_roster += 1

        if errors:
            msg = _("Import finished with errors:\n\n- %s") % ("\n- ".join(errors[:15]))
            if len(errors) > 15:
                msg += _("\n\n(+%s more)") % (len(errors) - 15)
            raise UserError(msg)

        message = _(
            "Roster import completed.\n"
            "- New contacts: %(cp)s\n"
            "- Updated contacts: %(up)s\n"
            "- New emergency contacts: %(ec)s\n"
            "- Roster lines created: %(cr)s\n"
            "- Already existed (skipped): %(sk)s"
        ) % {
            "cp": created_partners,
            "up": updated_partners,
            "ec": created_emergency_contacts,
            "cr": created_roster,
            "sk": skipped_existing,
        }

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import successful"),
                "message": message,
                "type": "success",
                "sticky": False,
                "next": {"type": "ir.actions.act_window_close"},
            },
        }
