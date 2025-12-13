# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
import base64
import openpyxl
from openpyxl import load_workbook
import io

class PartnerExcelImport(models.TransientModel):
    _name = 'lt.partner.excel.import'
    _description = 'LT Importar Participantes desde Excel'

    brigade_id = fields.Many2one('brigade', string='Brigada', required=True)
    import_file = fields.Binary(string='Archivo Excel', required=True)
    filename = fields.Char(string='Nombre Archivo')
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('ready', 'Listo para Importar'),
        ('done', 'Importado'),
    ], default='draft')

    preview_lines = fields.Text(string='Vista Previa')
    total_lines = fields.Integer(string='Total Líneas')
    imported_count = fields.Integer(string='Importados')
    error_count = fields.Integer(string='Errores')

    @api.onchange('import_file')
    def _onchange_import_file(self):
        self.preview_lines = self.total_lines = self.imported_count = self.error_count = 0
        self.state = 'draft'
        
        if self.import_file:
            try:
                wb = load_workbook(io.BytesIO(base64.b64decode(self.import_file)), read_only=True)
                ws = wb.active
                self.total_lines = ws.max_row - 1  # Sin header
                self.preview_lines = f"✅ Detectadas {self.total_lines} líneas de datos"
                self.state = 'ready'
                wb.close()
            except Exception as e:
                raise ValidationError(_('Error leyendo Excel: %s') % str(e))

    def action_import_excel(self):
        if not self.import_file:
            raise UserError(_('Debe seleccionar un archivo Excel'))

        wb = load_workbook(io.BytesIO(base64.b64decode(self.import_file)))
        ws = wb.active
        
        imported = errors = 0
        error_details = []

        # Mapa de columnas (A=1, B=2, etc.)
        col_map = {
            1: 'nombre',      # Columna A
            2: 'apellido',    # Columna B
            3: 'identificacion', # Columna C
            4: 'email',       # Columna D
            5: 'telefono',    # Columna E
            6: 'celular',     # Columna F
            7: 'direccion',   # Columna G
            8: 'ciudad',      # Columna H
            9: 'provincia',   # Columna I
        }

        for row_num in range(2, ws.max_row + 1):  # Empezar desde fila 2 (sin header)
            try:
                # Leer datos de la fila
                row_data = {}
                for col_num, field_name in col_map.items():
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[field_name] = str(cell_value).strip() if cell_value else ''

                nombre = row_data['nombre']
                apellido = row_data['apellido']
                
                if not nombre or not apellido:
                    errors += 1
                    error_details.append(f"Fila {row_num}: Nombre o apellido vacío")
                    continue

                partner_data = {
                    'name': f"{nombre} {apellido}".strip(),
                    'brigade_ids': [(4, self.brigade_id.id)],
                    'is_company': False,
                    'customer_rank': 1,
                }
                
                # Campos opcionales
                if row_data['identificacion']:
                    partner_data['LT_identificacion'] = row_data['identificacion']
                if row_data['email']:
                    partner_data['email'] = row_data['email']
                if row_data['telefono']:
                    partner_data['phone'] = row_data['telefono']
                if row_data['celular']:
                    partner_data['mobile'] = row_data['celular']
                if row_data['direccion']:
                    partner_data['street'] = row_data['direccion']

                self.env['res.partner'].create(partner_data)
                imported += 1
                
            except Exception as e:
                errors += 1
                error_details.append(f"Fila {row_num}: {str(e)}")

        self.write({
            'state': 'done',
            'imported_count': imported,
            'error_count': errors,
            'preview_lines': f"✅ {imported} creados, {errors} errores",
        })
        wb.close()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación Exitosa'),
                'message': f'{imported} participantes importados!\n{errors} errores encontrados.',
                'type': 'success' if errors == 0 else 'warning',
                'sticky': True,
            }
        }
